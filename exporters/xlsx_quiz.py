"""퀴즈 XLSX 백업."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = ["번호", "유형", "영역", "문항", "보기1", "보기2", "보기3", "보기4", "정답", "해설"]


def write(questions: list[dict], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "퀴즈"

    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    bold = Font(bold=True)
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, q in enumerate(questions, start=2):
        choices = (list(q.get("choices") or []) + ["", "", "", ""])[:4]
        ws.cell(row=row, column=1, value=q.get("no"))
        ws.cell(row=row, column=2, value=q.get("type"))
        ws.cell(row=row, column=3, value=q.get("category"))
        ws.cell(row=row, column=4, value=q.get("question"))
        for i, c in enumerate(choices, start=5):
            ws.cell(row=row, column=i, value=c)
        ws.cell(row=row, column=9, value=q.get("answer"))
        ws.cell(row=row, column=10, value=q.get("explanation"))

    widths = [6, 8, 14, 50, 24, 24, 24, 24, 8, 40]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    wb.save(path)
    return path
