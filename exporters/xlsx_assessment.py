"""학습평가 XLSX 백업 — 작업자가 편집·재활용하기 좋게 컬럼화."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = ["번호", "유형", "난이도", "문항", "보기1", "보기2", "보기3", "보기4", "정답", "해설", "근거"]


def write(questions: list[dict], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "학습평가"

    # 헤더 행
    header_fill = PatternFill(start_color="F4D6E5", end_color="F4D6E5", fill_type="solid")
    bold = Font(bold=True)
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터
    for row, q in enumerate(questions, start=2):
        choices = q.get("choices") or ["", "", "", ""]
        choices = (list(choices) + ["", "", "", ""])[:4]
        ws.cell(row=row, column=1, value=q.get("no"))
        ws.cell(row=row, column=2, value=q.get("type"))
        ws.cell(row=row, column=3, value=q.get("difficulty"))
        ws.cell(row=row, column=4, value=q.get("question"))
        for i, c in enumerate(choices, start=5):
            ws.cell(row=row, column=i, value=c)
        ws.cell(row=row, column=9, value=q.get("answer"))
        ws.cell(row=row, column=10, value=q.get("explanation"))
        ws.cell(row=row, column=11, value=q.get("source_page"))

    # 컬럼 폭
    widths = [6, 8, 8, 60, 30, 30, 30, 30, 8, 60, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    wb.save(path)
    return path
